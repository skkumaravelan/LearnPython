from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def format_all_headers(excel_path):
    """
    Format the header row (row 1) of all sheets in an Excel file.
    - Bold text
    - Yellow background
    - Calibri font, size 12
    - Center alignment

    Args:
        excel_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(excel_path)

    # Define header styling
    header_font = Font(name='Calibri', size=12, bold=True)
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply formatting to all cells in row 1 (header row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    wb.save(excel_path)
    print("  ✓ Headers formatted (bold, yellow, Calibri 12)")


def wrap_text_all_headers(excel_path):
    """
    Enable wrap text for all header rows (row 1) in all sheets.

    Args:
        excel_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(excel_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply wrap text to all cells in row 1 (header row)
        for cell in ws[1]:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal if cell.alignment else 'center',
                vertical=cell.alignment.vertical if cell.alignment else 'center',
                wrapText=True
            )

    wb.save(excel_path)
    print("  ✓ Wrap text enabled for headers")


def format_content_font_all_sheets(excel_path, font_name='Calibri', font_size=10):
    """
    Format the content (all rows except header row 1) in all sheets.
    Sets font to Calibri size 10 by default.

    Args:
        excel_path (str): Path to the Excel file to format.
        font_name (str): Font name (default: 'Calibri').
        font_size (int): Font size (default: 10).
    """
    wb = load_workbook(excel_path)

    # Define content font
    content_font = Font(name=font_name, size=font_size)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply font to all data rows (starting from row 2, skipping header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = content_font

    wb.save(excel_path)
    print(f"  ✓ Content font set to {font_name} {font_size}")


def auto_resize_all_columns(excel_path):
    """
    Auto-resize all columns in all sheets of an Excel file to fit content width.

    Args:
        excel_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(excel_path)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Auto-resize each column based on content
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set column width (add extra padding for readability)
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)
    print("  ✓ Columns auto-resized")


def set_max_column_width(excel_path, column_name='Event_Details', max_width=50):
    """
    Set a maximum width for a specific column across all sheets.
    Useful for columns with long text content to prevent excessive stretching.

    Args:
        excel_path (str): Path to the Excel file to format.
        column_name (str): Name of the column to limit (default: 'Event_Details').
        max_width (int): Maximum width in Excel units (default: 50).
    """
    wb = load_workbook(excel_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the column letter for the target column name
        column_letter = None
        for cell in ws[1]:
            if cell.value == column_name:
                column_letter = cell.column_letter
                break

        # If column found, set maximum width
        if column_letter:
            current_width = ws.column_dimensions[column_letter].width
            if current_width and current_width > max_width:
                ws.column_dimensions[column_letter].width = max_width

    wb.save(excel_path)
    print(f"  ✓ Column '{column_name}' width limited to {max_width}")


def enable_autofilter_all_sheets(excel_path):
    """
    Enable AutoFilter (dropdown arrows) for the header row in all sheets.

    Args:
        excel_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(excel_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Enable AutoFilter on the entire data range starting from row 1
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

    wb.save(excel_path)
    print("  ✓ AutoFilter enabled")


def freeze_top_row_all_sheets(excel_path):
    """
    Freeze the top row (header row) in all sheets so it stays visible when scrolling vertically.

    Args:
        excel_path (str): Path to the Excel file to format.
    """
    wb = load_workbook(excel_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Freeze the top row by setting freeze_panes to A2
        ws.freeze_panes = 'A2'

    wb.save(excel_path)
    print("  ✓ Top row frozen")


def rename_sheets(excel_path, sheet_name_mapping):
    """
    Rename sheets in the Excel file according to the provided mapping.

    Args:
        excel_path (str): Path to the Excel file.
        sheet_name_mapping (dict): Dictionary mapping old names to new names.
                                   Example: {'Sheet1': 'All Logs', 'Sheet2': 'Unique Codes'}
    """
    wb = load_workbook(excel_path)

    # Rename sheets according to mapping
    for old_name, new_name in sheet_name_mapping.items():
        if old_name in wb.sheetnames:
            ws = wb[old_name]
            ws.title = new_name
            print(f"  ✓ Renamed '{old_name}' → '{new_name}'")
        else:
            print(f"  ⚠ Sheet '{old_name}' not found, skipping")

    wb.save(excel_path)
    print("  ✓ Sheet names updated")
