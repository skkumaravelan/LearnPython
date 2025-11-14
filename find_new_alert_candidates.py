import pandas as pd
from openpyxl import load_workbook


def append_new_alerts_as_sheet5(existing_excel_path, configured_logs_path):
    # Load existing Excel file and Sheet2 with unique AMH log codes
    df_all_logs = pd.read_excel(existing_excel_path, sheet_name="Sheet2")

    # Find column with log codes
    col_candidates = [col for col in df_all_logs.columns if "code" in col.lower() or "log" in col.lower()]
    if not col_candidates:
        raise Exception('No suitable log code column found in Sheet2')
    log_code_col = col_candidates[0]
    unique_log_codes = set(df_all_logs[log_code_col].astype(str).str.strip())

    # Load already configured log codes from text file
    with open(configured_logs_path, 'r') as f:
        configured_log_codes = set(line.strip() for line in f if line.strip())

    # Determine potential new alerts
    potential_new_alerts = sorted(list(unique_log_codes - configured_log_codes))
    output_df = pd.DataFrame({"PotentialNewLogCodes": potential_new_alerts})

    # Load workbook directly using openpyxl
    book = load_workbook(existing_excel_path)
    # Remove existing Sheet5 if present to avoid duplicates or conflicts
    if "Sheet5" in book.sheetnames:
        std = book["Sheet5"]
        book.remove(std)
    # Save workbook to not lose sheets before use with ExcelWriter
    book.save(existing_excel_path)

    # Now use pandas ExcelWriter without setting 'book' attribute
    with pd.ExcelWriter(existing_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        output_df.to_excel(writer, sheet_name="Sheet5", index=False)

    return existing_excel_path, len(potential_new_alerts)
