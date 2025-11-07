from bs4 import BeautifulSoup
from openpyxl import Workbook

from bs4 import BeautifulSoup
from openpyxl import load_workbook


def read_swift_guide_and_write_to_excel(html_file_path, input_excel_path):
    """
    Reads event codes from Sheet2 column A (from A2 down to last non-empty),
    extracts details from HTML file, and writes results to Sheet3 of same Excel file.

    Args:
        html_file_path (str): Path to the HTML guide file.
        input_excel_path (str): Path to the Excel file with event codes and output target.
    """
    # Load workbook and select Sheet1 for reading event codes
    wb = load_workbook(input_excel_path)
    ws1 = wb["Sheet2"]  # Adjust if your sheet name differs

    # Read event codes from column A starting at row 2 until blank or None
    event_codes = []
    row = 2
    while True:
        cell_value = ws1[f"A{row}"].value
        if cell_value is None or str(cell_value).strip() == "":
            break
        event_codes.append(str(cell_value).strip())
        row += 1

    # Parse the HTML file to build lookup dictionary
    with open(html_file_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    code_map = {}
    for table in soup.find_all("table"):
        log_code = None
        log_message = None
        description = None
        for tr in table.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) >= 2:
                key = tds[0].get_text(strip=True)
                value = tds[1].get_text(strip=True)
                if "AMH_Event_Log_Code" in key:
                    log_code = value
                elif "Log Message" in key:
                    log_message = value
                elif "Description" in key:
                    description = value
        if log_code:
            code_map[log_code] = (log_message, description)

    # Gather output data for found codes
    output_data = []
    placeholder = "Data not available in Official SWIFT Log Guide"

    for code in event_codes:
        # Clean the code by removing brackets for lookup
        clean_code = code.replace("[", "").replace("]", "").strip()

        # Get message and description from code_map, use placeholder if not found
        if clean_code in code_map:
            message, desc = code_map[clean_code]
            # If either message or description is empty/missing, use placeholder
            message = message if message else placeholder
            desc = desc if desc else placeholder
        else:
            # Code not found in HTML at all
            message = placeholder
            desc = placeholder

        output_data.append((code, message, desc))

    # Write data to Sheet3 (overwrite if exists)
    if "Sheet3" in wb.sheetnames:
        ws3 = wb["Sheet3"]
        wb.remove(ws3)
    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["AMH_Event_Log_Code", "Log Message", "Description"])
    for row in output_data:
        ws3.append(row)

    # Save workbook
    wb.save(input_excel_path)
    print(f"Wrote {len(output_data)} Log Code details to Sheet3 of {input_excel_path}")


def read_swift_guide_and_write_to_excel_bkp(html_file_path, log_code_txt_path, output_excel_path):
    # Read log codes from txt and clean brackets
    with open(log_code_txt_path, "r", encoding="utf-8") as f:
        event_codes = [line.strip().replace("[", "").replace("]", "") for line in f if line.strip()]

    # 2. Parse the HTML just once into a dict for fast lookup
    with open(html_file_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    code_map = {}
    for table in soup.find_all("table"):
        log_code = None
        log_message = None
        description = None
        for tr in table.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) >= 2:
                key = tds[0].get_text(strip=True)
                value = tds[1].get_text(strip=True)
                if "AMH_Event_Log_Code" in key:
                    log_code = value
                elif "Log Message" in key:
                    log_message = value
                elif "Description" in key:
                    description = value
        if log_code:
            code_map[log_code] = (log_message, description)

    # 3. Gather output rows for the codes in txt
    output_data = []
    for code in event_codes:
        message, desc = code_map.get(code, ("", ""))
        output_data.append((code, message, desc))

    # 4. Write to Excel in one go
    wb = Workbook()
    ws = wb.active
    ws.title = "LogCodeDetails"
    ws.append(["AMH_Event_Log_Code", "Log Message", "Description"])
    for row in output_data:
        ws.append(row)
    wb.save(output_excel_path)
    print(f"Wrote {len(output_data)} Log Code details to {output_excel_path}")

# Usage:
# extract_log_details_to_excel(
#     "AMHApplicationLogGuide.html",
#     "codes.txt",
#     "LogCodeMessageDescriptions.xlsx"
# )
