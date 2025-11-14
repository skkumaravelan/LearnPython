import pandas as pd
from openpyxl import load_workbook

def create_splunk_candidates_sheet(excel_path, configured_txt_path):
    # Load Sheet2 and Sheet3
    sheet2 = pd.read_excel(excel_path, sheet_name='Sheet2')
    sheet3 = pd.read_excel(excel_path, sheet_name='Sheet3')
    # Merge based on event code
    merged = pd.merge(sheet2, sheet3, left_on='AMH_Event_Log_Code', right_on='AMH_Event_Log_Code', how='inner')
    # Add Splunk manual columns
    for col in ['Priority', 'Impact', 'Urgency', 'Threshold']:
        if col not in merged.columns:
            merged[col] = ''  # Placeholder for Splunk team to fill manually
    # Filter out any event already configured in Splunk
    with open(configured_txt_path) as f:
        configured_log_codes = set(line.strip() for line in f if line.strip())
    candidates = merged[~merged['AMH_Event_Log_Code'].astype(str).isin(configured_log_codes)]
    # Set the preferred column order for Splunk
    col_order = ['AMH_Event_Log_Code', 'Status', 'Priority', 'Impact', 'Urgency', 'Threshold', 'Log Message', 'Description']
    # Only keep columns present in the DataFrame to avoid key errors
    col_order_final = [c for c in col_order if c in candidates.columns]
    candidates = candidates[col_order_final]
    # Remove existing Sheet4 if needed
    book = load_workbook(excel_path)
    if 'Sheet4' in book.sheetnames:
        book.remove(book['Sheet4'])
        book.save(excel_path)
    # Write candidates as new Sheet4
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        candidates.to_excel(writer, sheet_name='Sheet4', index=False)
    print("Final candidate list written to Sheet4! This is your Splunk-ready sheet.")

# Example call
# create_splunk_candidates_sheet(r"S:\LearnPython\outputs\all_logs.xlsx", r"S:\LearnPython\outputs\Already_configured_logcodes_in_amh.txt")
